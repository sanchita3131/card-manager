"""
Excel writer — creates and manages a single .xlsx workbook in memory.
No Google Cloud, no API keys, no setup.
"""

import logging
from typing import Dict, Optional, List
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

HEADERS = [
    "Company Name",
    "Card Holder Name",
    "Position",
    "Contact Number",
    "Email Address",
    "Timestamp",
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook() -> Workbook:
    """
    Create a new Excel workbook with headers.

    Returns:
        openpyxl Workbook object (ready to append rows)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Cards"

    # Style the header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = THIN_BORDER

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 22

    # Freeze the header row
    ws.freeze_panes = "A2"

    return wb


def append_card(wb: Workbook, card_data: Dict[str, str]) -> bool:
    """
    Append a card entry to the workbook.

    Args:
        wb: openpyxl Workbook object
        card_data: Dict with keys: company, name, position, phone, email

    Returns:
        True always (failure raises exception)
    """
    try:
        ws = wb.active
        row = [
            card_data.get("company", "null"),
            card_data.get("name", "null"),
            card_data.get("position", "null"),
            card_data.get("phone", "null"),
            card_data.get("email", "null"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

        # Apply borders to the new row
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = THIN_BORDER

        logger.info(f"Card appended to Excel: {card_data.get('name')}")
        return True

    except Exception as e:
        logger.error(f"Failed to append to workbook: {e}")
        return False


def get_all_cards(wb: Workbook) -> Optional[List[Dict[str, str]]]:
    """
    Retrieve all card entries from the workbook.

    Args:
        wb: openpyxl Workbook object

    Returns:
        List of dicts, or None on error
    """
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        records = []
        for row in rows:
            if any(cell is not None for cell in row):
                records.append({
                    "Company Name": row[0] or "null",
                    "Card Holder Name": row[1] or "null",
                    "Position": row[2] or "null",
                    "Contact Number": row[3] or "null",
                    "Email Address": row[4] or "null",
                    "Timestamp": str(row[5]) if row[5] else "",
                })
        return records

    except Exception as e:
        logger.error(f"Failed to read workbook: {e}")
        return None


def to_bytes(wb: Workbook) -> BytesIO:
    """
    Serialize workbook to a BytesIO object for download.

    Args:
        wb: openpyxl Workbook object

    Returns:
        BytesIO containing the .xlsx data
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
